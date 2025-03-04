import base64
import datetime
import os
from collections import Counter
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from pprint import pformat
from typing import Union, List
from urllib.parse import quote

import jieba
import matplotlib
from loguru import logger
from matplotlib.font_manager import FontProperties
from optbinning.binning.binning_statistics import MulticlassBinningTable
from wordcloud import WordCloud

asset_home = Path(__file__).parent / 'assets'
font_path = asset_home / 'msyhl.ttc'
font_prop = FontProperties(fname=font_path)
matplotlib.rcParams['font.family'] = font_prop.get_name()
matplotlib.rcParams['font.sans-serif'] = [font_prop.get_name(), 'SimHei']
matplotlib.rcParams['axes.unicode_minus'] = False

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
from optbinning import BinningProcess, ContinuousOptimalBinning, OptimalBinning, MulticlassOptimalBinning
from optbinning.binning.auto_monotonic import type_of_monotonic_trend
from scipy.stats import spearmanr, kendalltau
import ydata_profiling
from ydata_profiling import ProfileReport
from ydata_profiling.config import Settings
from ydata_profiling.model.typeset import ProfilingTypeSet
from bibidataprofile.optbinning_plot import plot_optbinning



# def _plot_word_cloud(
#         config: Settings,
#         series: Union[pd.Series, List[pd.Series]],
#         figsize: tuple = (6, 4),
# ) -> plt.Figure:
#     if not isinstance(series, list):
#         series = [series]
#     plot = plt.figure(figsize=figsize)
#     for i, series_data in enumerate(series):
#         word_dict = series_data.to_dict()
#         wordcloud = WordCloud(
#             font_path=font_path,
#             background_color="white", random_state=123, width=300, height=200, scale=2
#         ).generate_from_frequencies(word_dict)
#
#         ax = plot.add_subplot(1, len(series), i + 1)
#         ax.imshow(wordcloud)
#         ax.axis("off")
#
#     return plot


# ydata_profiling.visualisation.plot._plot_word_cloud = _plot_word_cloud


def _bin_str_label_format(bin_str, max_length=27):
    _bin_str = []
    for bs in bin_str:
        label = str(bs)
        if len(label) > max_length:
            label = label[:max_length] + '...'
        _bin_str.append(label)

    return _bin_str


@dataclass(frozen=True)
class DataTypes:
    Id: str = 'Id'
    Numeric: str = 'Numeric'
    Text: str = 'Text'
    Datetime: str = 'Datetime'
    Categorical: str = 'Categorical'
    Unsupported: str = 'Unsupported'


@dataclass(frozen=True)
class EstimatorType:
    Classification = "Classification"
    Regression = "Regression"


HTML_TEMPLATE = """
<!DOCTYPE HTML>
<html>
    <meta charset="utf-8" />
    <head>
    <title>BibiDataProfile Report</title>
    <style>
        body, p, td, pre {
            font-size: 14pt; }

        body {
            background-color:lightyellow;
            counter-reset: section subsection figure;}

        h1 {
            text-align:center }

        h2 {
            counter-increment: section; }

        h3 {
            counter-increment: subsection; }

        h4 {
            counter-increment: figure; }

        h2:before {
            content:counter(section) ". "; }

        h3:before {
            content:counter(section) "." counter(subsection) " "; }

        h4:before{
            content:"Fig. " counter(section) "." counter(figure) " - ";}

        .reset {
            counter-reset:subsection figure;}

        .nav-bar{
            list-style-type:none;
            margin:0;
            padding:0; }

        img {
            margin-left: 10pt;
            border: thin solid gray; }

        .sample {
            background-color: black;
            font-size: 12pt;
            color:white;
            padding: 4pt;
            margin-left: 10pt;
            margin-right: 30pt;
            display: block;
            white-space: pre-line;}

        .code {
            padding: 4pt;
            font-size: 12pt;
            background-color: whitesmoke;
            border: thin dotted gray;
            margin-left: 10pt;
            margin-right: 30pt;
            display: block;
            white-space: pre-line;}

        .author {
            font-size: 16pt;
            text-align:center;}

        /*  dataframe tables  */
        table.dataframe {
            margin: 8px;
            border-width: 1px;
            border-color: gray;
            border-collapse: collapse; }

        table.dataframe th {
            border-width: 1px;
            padding: 8px 8px 8px 8px;
            border-style: solid;
            border-color: gray;
            background-color: lightyellow; }

        table.dataframe td {
            text-align:center;
            border-width: 1px;
            padding: 8px 8px 8px 8px;
            border-style:solid;
            border-color: gray;
            background-color: #ffffff; }

        /*  grid tables  */
        table.g {
            margin: 8px;
            border-width: 1px;
            border-color: gray;
            border-collapse: collapse; }

        table.g th {
            border-width: 1px;
            padding: 8px 8px 8px 8px;
            border-style: solid;
            border-color: gray;
            background-color: lightyellow; }

        table.g td {
            text-align:center;
            border-width: 1px;
            padding: 8px 8px 8px 8px;
            border-style:solid;
            border-color: gray;
            background-color: #ffffff; }

        /*  vertical tables  */
        table.v {
            margin: 8px;
            border-collapse: collapse; }

        table.v thead {
            border-right: 10px solid;
            border-left: 10px solid;
            background: #fc9;
            border-color: #fc9;}

        table.v th {
            padding: 4px 20px 4px 20px;
            border: 1px #fff solid; }

        table.v tbody{
            text-align:left;
            border-right: 10px solid;
            border-left: 10px solid;
            border-color: #adf;
            background: #adf;
            vertical-align: top; }

        table.v td {
            padding:  4px 20px 4px 20px;
            border: 1px #fff solid; }

        /*  horizontal tables  */
        table.h {
            margin: 8px;
            border-collapse:collapse; }

        table.h th {
            padding:  4px 40px 4px 40px;
            border:1px solid #98bf21;
            background-color:#A7C942;
            color:#fff; }

        table.h td {
            text-align:center;
            border:1px solid #98bf21;
            padding: 4px 40px 4px 40px; }

        table.h tr.alt td {
            color:#000;
            background-color:#EAF2D3; }
        /**
            http://learnlayout.com/inline-block.html
            http://www.w3schools.com/cssref/css_colornames.asp
        **/

        div.link {
            display: inline-block;
            width: 80px;
            height: 100px;
            position: relative;
            text-align:center;
            vertical-align: middle; }

        div.black {
            border-color: Black ;
            border-width: 2px; }

        div.black span.h {
            background-color: Black       ;
            color: Snow ;  }

        div.white {
            border-color: Lavender  ;
            border-width: 2px; }

        div.white span.h {
            background-color: Snow       ;
            color: Black ; }

        div.red {  border-color: HotPink }

        div.red span.h {
            background-color: HotPink;
            color: LightCyan;         }

        div.yellow { border-color: GoldenRod  }

        div.yellow span.h {
            background-color: GoldenRod    ;
            color: Azure ;         }

        div.blue { border-color: DodgerBlue }

        div.blue span.h {
            background-color: DodgerBlue     ;
            color: Snow ;  }

        div.green { border-color: ForestGreen }

        div.green span.h {
            background-color: MediumSeaGreen    ;
            color: Ivory   ; }

        div.box {
            display: inline-block;
            width: 200px;
            height: 100px;
            margin: 8px;
                border-style:solid;
            border-width: 1.5px;
            position: relative;
            text-align:center;
            vertical-align: middle;  }

        div.box span.h {
            top: 0;
            left: 0;
            right: 0;
            text-align:left;
            padding: 1px 2px 2px 1px;
            position: absolute;
            font-weight:bold;
            line-height: 1em;  }

        </style></head>
<body>
    <a name="_top_" />
    <nav class="nav-bar">
        <a href="#binning-analysis">Binning Analysis</a>
	</nav>
    <h1>BibiDataProfile Report</h1>
    <p class="author">
 Version <var>1</var>, <ruby>bibidataprofile <rt> Chunqi SHI</rt></ruby>,  $today</p>


<a name="binning-analysis" />
<h2>Binning Analysis  <a href="#_top_" style="text-decoration: none;">^</a></h2>

$binning_analysis

<footer><hr/>
<p>
    Progress:
    <progress value="100" max="100"></progress></p>
<p>

</body></html>
"""

__extmap__ = {
    ".gif": "image/gif",
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".svg": "image/svg+xml",
}


def base64_image(image: bytes, mime_type: str = __extmap__[".png"]) -> str:
    """Encode the image for an URL using base64

    Args:
        image: the image
        mime_type: the mime type

    Returns:
        A string starting with "data:{mime_type};base64,"
    """
    base64_data = base64.b64encode(image)
    image_data = quote(base64_data)
    return f'<img src="data:{mime_type};base64,{image_data}"/>'


def image_to_base64(image_path) -> str:
    """image to html base64 string
    ...

    Args
    ----------
    image_path : str
       image path

    Returns
    ----------
       str: encoding image stri
    """
    import base64

    (nam, ext) = os.path.splitext(image_path)
    nam = os.path.basename(image_path)
    with open(image_path, "rb") as image_file:
        encoded_string = base64.b64encode(image_file.read())
        encoded_string = str(encoded_string, encoding="utf-8")
    base64_string = (
            '<img src="data:'
            + __extmap__[ext.lower()]
            + ";base64,"
            + encoded_string
            + '" alt="'
            + nam
            + '"/>'
    )
    return base64_string


def generate_factor_analysis_html(
        binning_analysis=""
):
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    from string import Template

    template = Template(HTML_TEMPLATE)
    string = template.substitute(
        binning_analysis=binning_analysis,
        today=today,
    )
    return string


def generate_binning_analysis(estimator_type, factors, tables, images) -> str:
    html_text = ""
    for i, factor in enumerate(factors):
        html_text += f"<h3>Factor {i + 1}th: {factor}</h3>\n"
        html_text += (
            f"<p><b>Table {i + 1}th: {factor} binning table dataframe</b></p>\n"
        )
        html_text += tables[i].to_html()
        if estimator_type == EstimatorType.Regression:
            img_base64 = images[i]
            html_text += f"\n<figure>{img_base64}<figcaption><h4>{i + 1}th: {factor} binning table mean</h4></figcaption></figure>\n"
        if estimator_type == EstimatorType.Classification:
            if len(images) > len(factors):
                img_base64 = images[2 * i + 0]
                html_text += f"\n<figure>{img_base64}<figcaption><h4>{i + 1}th: {factor} binning table woe</h4></figcaption></figure>\n"
                img_base64 = images[2 * i + 1]
                html_text += f"\n<figure>{img_base64}<figcaption><h4>{i + 1}th: {factor} binning table event rate</h4></figcaption></figure>\n"
            else:
                img_base64 = images[i]
                html_text += f"\n<figure>{img_base64}<figcaption><h4>{i + 1}th: {factor} binning table event rate</h4></figcaption></figure>\n"
    return html_text


def binning_analysis_plot(
        X,
        y,
        estimator_type,
        type_variables=None,
        binning_fit_params=None,
        html_path="binning_analysis.html"
) -> (list, list, str):
    """binning analysis for each factors

    Parameters
    ----------
        X : pandas.DataFrame
            training X factors in DataFrame
        y : pandas.DataFrame
            training y target in DataFrame
        html_path : str
            html path

    Returns
    -------
        3 tuples (tables, images, html_text) are returned.
        tables : list[pandas.DataFrame]
            binning table for each factors
        images : list[str]
            images such as mean plot for regression, woe, event rate plot for classification for each factor
        html_text : str
            generated html

    """

    if type_variables is None:
        type_variables = get_data_types(X)
    # variable_types = reverse_mapping(type_variables)

    factors = [f for f in X.columns.tolist() if f in type_variables.get('Numeric', [])]
    tables = []
    images = []
    if estimator_type == EstimatorType.Regression:
        for i, factor in enumerate(factors):
            logger.info(f'factor={factor}')
            optb = ContinuousOptimalBinning(
                name=factor,
                prebinning_method="cart",
                max_n_prebins=60,
                min_prebin_size=0.01,
                max_bin_size=0.2,
                monotonic_trend="auto",
            )
            # Input X contains infinity or a value too large for dtype('float32').
            logger.info(f'X[factor]=[{len(X[factor]), len(X[factor].dropna())}], y=[{len(y), len(y.dropna())}]')
            X_dropna, y_dropna = drop_null_xy(X[factor], y)
            logger.info(f'X_dropna=[{len(X_dropna)}], y_dropna=[{len(y_dropna)}]')
            X_dropna = X_dropna.astype(float)
            optb.fit(X_dropna, y_dropna)
            binning_table = optb.binning_table
            df_table = binning_table.build()
            tables.append(df_table)
            image_bytes = BytesIO()
            plot_optbinning(
                binning_table,
                metric='mean',
                savefig=image_bytes
            )
            images.append(base64_image(image_bytes.getvalue()))

    if estimator_type == EstimatorType.Classification:
        ismulticlass = y.nunique() > 2
        for i, factor in enumerate(factors):
            logger.info(f'factor={factor}')
            logger.info(f'X[factor]=[{len(X[factor]), len(X[factor].dropna())}], y=[{len(y), len(y.dropna())}]')
            X_dropna, y_dropna = drop_null_xy(X[factor], y)
            logger.info(f'X_dropna=[{len(X_dropna)}], y_dropna=[{len(y_dropna)}]')
            X_dropna = X_dropna.astype(float)
            if ismulticlass:
                optb = MulticlassOptimalBinning(
                    name=factor,
                    # prebinning_method="cart",
                    # max_n_prebins=60,
                    # min_prebin_size=0.01,
                    # max_bin_size=0.2,
                    # monotonic_trend="auto",
                )
                optb.fit(X_dropna, y_dropna)
                binning_table = optb.binning_table
                df_table = binning_table.build()
                tables.append(df_table)
                image_bytes = BytesIO()
                plot_optbinning(
                    binning_table,
                    savefig=image_bytes
                )
                images.append(base64_image(image_bytes.getvalue()))

            else:
                optb = OptimalBinning(
                    name=factor,
                    # prebinning_method="cart",
                    # max_n_prebins=60,
                    # min_prebin_size=0.01,
                    # max_bin_size=0.2,
                    # monotonic_trend="auto",
                )
                optb.fit(X_dropna, y_dropna)
                binning_table = optb.binning_table
                df_table = binning_table.build()
                tables.append(df_table)
                image_bytes = BytesIO()
                plot_optbinning(
                    binning_table,
                    metric="woe",
                    savefig=image_bytes
                )
                images.append(base64_image(image_bytes.getvalue()))
                image_bytes = BytesIO()
                plot_optbinning(
                    binning_table,
                    metric="event_rate",
                    savefig=image_bytes,
                )
                images.append(base64_image(image_bytes.getvalue()))

    html_text = generate_binning_analysis(
        estimator_type, factors, tables, images
    )
    html_text_with_header = generate_factor_analysis_html(
        binning_analysis=html_text
    )
    if html_path is not None:
        with open(html_path, "w", encoding="utf-8") as html_fp:
            html_fp.write(html_text_with_header)
    return tables, images, html_text


# def detect_csv_encoding(file_path):
#     import chardet
#     with open(file_path, 'rb') as f:
#         result = chardet.detect(f.read(1024))
#     return result['encoding']


def read_txt_file(data_file, encoding):
    df = None
    suffix = str(data_file).split('.')[-1].lower()
    if suffix in ['csv', 'tsv']:
        df = pd.read_csv(data_file, encoding=encoding)
    elif suffix in ['json']:
        df = pd.read_json(data_file, encoding=encoding)
    elif suffix in ['xml']:
        df = pd.read_xml(data_file, encoding=encoding)
    return df


def read_large_excel(io, sheet_name="Sheet1"):
    import openpyxl
    workbook = openpyxl.load_workbook(io, read_only=True)
    worksheet = workbook[sheet_name]
    rows = []
    for row in worksheet.iter_rows(values_only=True):
        rows.append(row)
    dataDF = pd.DataFrame(rows[1:], columns=rows[0])
    return dataDF


# def read_excel(data_file):
#     if Path(data_file).stat().st_size > 64 * 1024 * 1024:
#         try:
#             df = read_large_excel(data_file)
#             return df
#         except:
#             pass
#     df = pd.read_excel(data_file)
#     return df

def read_excel(data_file):
    try:
        df = pd.read_excel(data_file, engine="calamine")
        return df
    except:
        df = pd.read_excel(data_file)
    return df


def read_data_file(data_file, encodings=None):
    if not Path(data_file).exists():
        return None
    if not Path(data_file).is_file():
        return None
    if encodings is None:
        encodings = ['utf8', 'GB18030', 'gbk', 'big5']
    df = None
    suffix = str(data_file).split('.')[-1].lower()
    if suffix in ['csv', 'tsv', 'json', 'xml']:
        for encoding in encodings:
            try:
                df = read_txt_file(data_file, encoding)
                break
            except:
                pass
    elif suffix in ['xls', 'xlsx']:
        df = read_excel(data_file)
    elif suffix in ['hdf', 'h5']:
        df = pd.read_hdf(data_file)
    return df


def to_color_excel(df_data: pd.DataFrame, xlsx_path, write_index=True):
    """
    colorize pandas.DataFrame according to quartile ratio
    ...

    Args
    ----------
    df_data : pandas.DataFrame
        data frame
    xlsx_path : str
        saving path of excel

    """
    score_titles = ["score", "scores", "value", "values", "打分", "分值", "分数", "评分"]
    quartiles = [0, 0.25, 0.5, 0.75, 1]
    colors_light = ["00ffd4", "98f5ff", "fed8b1", "ff7f7d"]
    colors_dark = ["40E0D0", "6495ED", "FF7F50", "FC3468"]
    try:
        if write_index:
            df_data = df_data.copy(deep=True).reset_index()
        else:
            df_data = df_data.copy(deep=True).reset_index(drop=True)
        wb = Workbook()
        ws = wb.active

        for c_idx, column in enumerate(df_data.columns):
            ws.cell(row=1, column=c_idx + 1, value=column)

        for r_idx, row in df_data.iterrows():
            for c_idx, val in enumerate(row):
                try:
                    ws.cell(row=r_idx + 2, column=c_idx + 1, value=val)
                except:
                    ws.cell(row=r_idx + 2, column=c_idx + 1, value=str(val))
        for c_idx, column in enumerate(df_data.columns):
            try:
                pd.to_numeric(df_data[column])
                colors = colors_light
            except:
                continue
            if any(subs in str(column).lower() for subs in score_titles):
                colors = colors_dark
            values = pd.to_numeric(df_data[column])
            quartile_values = [values.quantile(q) for q in quartiles]
            for r_idx, val in enumerate(values):
                cell = ws.cell(row=r_idx + 2, column=c_idx + 1)
                for i in range(len(quartiles) - 1):
                    if val <= quartile_values[i + 1]:
                        cell.fill = PatternFill(
                            start_color=colors[i],
                            end_color=colors[i],
                            fill_type="solid",
                        )
                        break
        wb.save(xlsx_path)
    except Exception as ex:
        logger.warning(
            f"to_color_excel() FAILED because of {ex}, using pandas DataFrame to_excel()"
        )
        df_data.to_excel(xlsx_path)


def is_series_numeric(series: pd.Series):
    try:
        pd.to_numeric(series)
        return True
    except:
        return False


def is_series_text(series: pd.Series,
                   MINIMUM_TEXT_CUTS_AS_TEXT=5,
                   MAXIMUM_UNIQUE_VALUES_AS_DICTIONARY=3
                   ):
    text_cuts_sr = series.dropna().apply(lambda x: list(jieba.cut(str(x))))
    number_text_cuts_sr = text_cuts_sr.apply(lambda x: len(x))
    word_counts = Counter()
    for cuts in text_cuts_sr:
        word_counts.update(cuts)
    if number_text_cuts_sr.max() <= MINIMUM_TEXT_CUTS_AS_TEXT:
        return False
    max_value_length = max([value for value in dict(word_counts).values()])
    if max_value_length <= MAXIMUM_UNIQUE_VALUES_AS_DICTIONARY:
        return False
    return True


def is_series_categorical(series: pd.Series,
                          MAXIMUM_UNIQUE_TEXT_AS_CATEGORICAL=20,
                          MAXIMUM_UNIQUE_NUMERIC_AS_CATEGORICAL=8,
                          CATEGORICAL_PERCENTAGE_THRESHOLD_DEFAULT=0.2):
    nunique = series.nunique()
    if is_series_numeric(series):
        if nunique > MAXIMUM_UNIQUE_NUMERIC_AS_CATEGORICAL:
            return False
        if not pd.api.types.is_integer_dtype(series):
            return False
    else:
        if nunique > MAXIMUM_UNIQUE_TEXT_AS_CATEGORICAL:
            return False
    if series.nunique() / len(series.dropna()) > CATEGORICAL_PERCENTAGE_THRESHOLD_DEFAULT:
        return False
    return True


def get_data_types(data_df: pd.DataFrame, variables=None):
    if variables is None:
        variables = data_df.columns

    unsupporteds = []
    categoricals = []
    datetimes = []
    numerics = []
    texts = []
    ids = []
    for variable in variables:
        # logger.debug(variable)
        # logger.debug(data_df[variable])
        # logger.debug(data_df[variable].nunique())
        if data_df[variable].nunique() < 1:
            unsupporteds.append(variable)
        elif isinstance(data_df[variable].dtype, pd.CategoricalDtype):
            categoricals.append(variable)
            try:
                pd.to_numeric(data_df[variable])
                numerics.append(variable)
            except:
                texts.append(variable)
        else:
            if is_series_categorical(data_df[variable]):
                categoricals.append(variable)
            elif pd.api.types.is_datetime64_any_dtype(data_df[variable]):
                datetimes.append(variable)
            elif pd.api.types.is_numeric_dtype(data_df[variable]):
                numerics.append(variable)
            else:
                try:
                    pd.to_numeric(data_df[variable])
                    numerics.append(variable)
                except:
                    try:
                        pd.to_datetime(data_df[variable])
                        datetimes.append(variable)
                    except:
                        if is_series_text(data_df[variable]):
                            texts.append(variable)
                        else:
                            ids.append(variable)

    return dict(DateTime=datetimes, Numeric=numerics,
                Text=texts, Categorical=categoricals,
                Id=ids, Unsupported=unsupporteds)


def replace_strings_in_file(input_path, output_path, string_map: dict, encoding='utf8'):
    try:
        input_str = Path(input_path).read_text(encoding=encoding)
        for key, val in string_map.items():
            input_str = input_str.replace(key, val)
        Path(output_path).write_text(input_str, encoding=encoding)
    except Exception as e:
        logger.warning(f'Exception: {e}')


def gen_data_profile(data, workdir=None, title='Data Profile', html_file='data_profile.html'):
    if workdir is None:
        workdir = '.'
    profile = ProfileReport(data, title=title, minimal=True)
    profile.config.plot.font_path = font_path
    html_path = Path(workdir) / html_file
    profile.to_file(html_path)
    string_map = {
        '<p class="text-body-secondary text-center">Report generated by <a href="https://ydata.ai/?utm_source=opensource&utm_medium=pandasprofiling&utm_campaign=report">YData</a>.</p>': '',
        '<p class="text-body-secondary text-end">Brought to you by <a href="https://ydata.ai/?utm_source=opensource&utm_medium=ydataprofiling&utm_campaign=report">YData</a></p>': '',
        '<a href=https://github.com/ydataai/ydata-profiling>ydata-profiling vv4.12.2</a>': '',
        '<meta name=viewport content="width=device-width, initial-scale=1, shrink-to-fit=no"><meta name=author content="YData and the open source community."><meta name=generator content="YData Profiling vv4.12.2"><meta name=url content=https://github.com/ydataai/ydata-profiling>': ''
    }
    replace_strings_in_file(html_path, html_path, string_map=string_map)
    return str(html_path.resolve())


def get_numeric_columns(data_df: pd.DataFrame, variables=None):
    if variables is None:
        variables = data_df.columns
    numeric_variables, non_numeric_variables = [], []
    for variable in variables:
        try:
            pd.to_numeric(data_df[variable])
            numeric_variables.append(variable)
        except:
            non_numeric_variables.append(variable)

    logger.info(f'variables : \n{pformat(variables)}')
    logger.info(f'numeric_variables : \n{pformat(numeric_variables)}')
    logger.info(f'non_numeric_variables : \n{pformat(non_numeric_variables)}')
    return numeric_variables, non_numeric_variables


def get_variable_types_simple(data_df):
    typeset = ProfilingTypeSet(Settings())
    variable_types = typeset.detect_type(data_df)
    return variable_types


def get_variable_types(data_df):
    '''
    "Boolean": render_algorithms.render_boolean,
    "Numeric": render_algorithms.render_real,
    "Complex": render_algorithms.render_complex,
    "Text": render_algorithms.render_text,
    "DateTime": render_algorithms.render_date,
    "Categorical": render_algorithms.render_categorical,
    '''
    typeset = ProfilingTypeSet(Settings())
    variable_type_map = typeset.infer_type(data_df)
    # logger.info(f'variable_type_map : \n{pformat(variable_type_map)}')
    type_variable_map = reverse_mapping(variable_type_map)
    # logger.info(f'type_variable_map : \n{pformat(type_variable_map)}')
    return type_variable_map


def reverse_mapping(original_dict):
    reversed_dict = {}
    first_value = next(iter(original_dict.values()))
    if isinstance(first_value, list):
        for data_type, variables in original_dict.items():
            for variable in variables:
                reversed_dict[str(variable)] = reversed_dict.get(str(variable), []) + [data_type]
    else:
        for variable, data_type in original_dict.items():
            reversed_dict[str(data_type)] = reversed_dict.get(str(data_type), []) + [variable]
    return reversed_dict


def drop_null_Xy(X: pd.DataFrame, y: pd.Series):
    concatenated = pd.concat([X, y], axis=1)
    concatenated = concatenated.replace([np.inf, -np.inf], np.nan).dropna()
    X_dropna = concatenated.iloc[:, :-1]
    y_dropna = concatenated.iloc[:, -1]
    return X_dropna, y_dropna


def drop_null_xy(x: pd.Series, y: pd.Series):
    concatenated = pd.concat([x, y], axis=1)
    concatenated = concatenated.replace([np.inf, -np.inf], np.nan).dropna()
    X_dropna = concatenated.iloc[:, 0]
    y_dropna = concatenated.iloc[:, 1]
    return X_dropna, y_dropna


def fill_null_by_types(X: pd.DataFrame, y: pd.Series, variable_types: dict):
    x_factors = X.columns.tolist()
    mask = ~y.replace([np.inf, -np.inf], np.nan).isnull()
    X_fillna = X[mask]
    y_fillna = y[mask]
    for x_factor in x_factors:
        types = variable_types[x_factor]
        if 'Numeric' in types:
            X_fillna[x_factor] = X_fillna[x_factor].replace([np.inf, -np.inf], np.nan).fillna(0)
        elif 'Text' in types:
            X_fillna[x_factor] = X_fillna[x_factor].replace([np.inf, -np.inf], np.nan).fillna('')
        elif 'DateTime' in types:
            X_fillna[x_factor] = X_fillna[x_factor].replace([np.inf, -np.inf], np.nan).fillna(pd.Timestamp("19700101"))
        else:
            X_fillna[x_factor] = X_fillna[x_factor].replace([np.inf, -np.inf], np.nan).fillna(0)

    return X_fillna, y_fillna


def text_preprocessing(X, text_features):
    X_preprocessed = X.copy()
    for feature in text_features:
        X_preprocessed[feature] = X[feature].apply(lambda x: ' '.join(jieba.cut(x)))
    return X_preprocessed


def fit_regression(
        df_data,
        X_factors,
        y_factor,
        xlsx_path=None,
        binning_fit_params=None,
        type_variables=None,
        verbose=False,
):
    """
    analyze factor importance of regression task

    Parameters
    ----------
        df_data : pandas.DataFrame
            training table in DataFrame
        X_factors : list[str]
            X factor variable names in df_data columns
        y_factor : str
            y factor variable name in df_data columns
        xlsx_path : str
            result DataFrame output as Excel
        binning_fit_params:dict
            binning fit parameters, see https://gnpalencia.org/optbinning/binning_process.html
        verbose : bool
            logs information output
    Returns
    -------
    df_factor_analysis : pandas.DataFrame
        result table in pandas DataFrame with one factor in each row, including.
    1. n_bins: see https://gnpalencia.org/optbinning/binning_tables.html
    2. rank_order_correlation(kendalltau): see https://docs.scipy.org/doc/scipy/reference/generated/scipy.stats.kendalltau.html
    3. rank_order_correlation(spearmanr): see https://docs.scipy.org/doc/scipy/reference/generated/scipy.stats.spearmanr.html
    4. min_value
    5. max_value
    6. bins: see https://gnpalencia.org/optbinning/binning_tables.html
    7. missing
    8. std
    9. woe: see https://gnpalencia.org/optbinning/binning_tables.html
    10. quality_score: see https://gnpalencia.org/optbinning/binning_tables.html
    11. feature_importance(RandomForestRegressor): see https://scikit-learn.org/stable/modules/permutation_importance.html
    12. feature_importance(XGBRegressor): see https://scikit-learn.org/stable/modules/permutation_importance.html


    """
    if type_variables is None:
        type_variables = get_data_types(df_data)
    variable_types = reverse_mapping(type_variables)
    X = df_data[X_factors]
    y = df_data[y_factor]
    if verbose:
        logger.info(f"regression factor analysis, with X[{X.shape}], y[{y.shape}]")
    X_fillna, y_fillna = fill_null_by_types(X, y, variable_types)
    y_fillna = y_fillna.astype("float")

    if verbose:
        logger.info(f"X_fillna = [{X_fillna.shape}], isnull={X_fillna.isnull().sum().sum()}")
        logger.info(f"y_fillna = [{y_fillna.shape}], isnull={y_fillna.isnull().sum()}")

    if binning_fit_params is None:
        binning_fit_params = {}
    binning_fit_params_all = dict(
        zip(X_factors, [binning_fit_params] * len(X_factors))
    )
    binning_process = BinningProcess(
        variable_names=X_factors, binning_fit_params=binning_fit_params_all, verbose=verbose
    )
    binning_process.fit(X_fillna, y_fillna * (1 + 1e-6))
    df_quality = binning_process.summary()
    trends = []
    mins = []
    maxs = []
    bins = []
    missings = []
    stds = []
    spearmans = []
    kendalltaus = []
    dtypes = []

    for idx, row in df_quality.iterrows():
        name = row["name"]
        dtype = row["dtype"]
        dtypes.append(dtype)
        if verbose:
            logger.info(f"analyze factor: {name}")
        binning_table = binning_process._binned_variables[name]._binning_table
        trend = type_of_monotonic_trend(binning_table._mean[:-2])
        trends.append(trend)
        mins.append(binning_table.min_x)
        maxs.append(binning_table.max_x)
        bins.append(list(binning_table._bin_str[:-2]))
        missings.append(round(df_data[name].isnull().mean(), 2))
        if dtype in ['numerical']:
            stds.append(df_data[name].std())
            spearman_score, spearman_p_value = spearmanr(X_fillna[name], y_fillna)
            spearmans.append(spearman_score)
            kendalltau_score, kendalltau_p_value = kendalltau(X_fillna[name], y_fillna)
            kendalltaus.append(kendalltau_score)
        else:
            stds.append(0)
            spearmans.append(0)
            kendalltaus.append(0)

    df_quality["dtype"] = dtypes
    df_quality["monotonic_trend"] = trends
    df_quality["min_value"] = mins
    df_quality["max_value"] = maxs
    df_quality["bins"] = bins
    df_quality["missing"] = missings
    df_quality["std"] = stds
    df_quality["rank_order_correlation(spearmanr)"] = spearmans
    df_quality["rank_order_correlation(kendalltau)"] = kendalltaus

    columns = [
        "name",
        "dtype",
        "n_bins",
        "monotonic_trend",
        "rank_order_correlation(kendalltau)",
        "rank_order_correlation(spearmanr)",
        "min_value",
        "max_value",
        "bins",
        "missing",
        "std",
        "woe",
        "quality_score",
    ]

    df_quality = df_quality[columns]
    df_quality = df_quality.sort_values(["quality_score"], ascending=False)
    df_factor_analysis = df_quality.copy(deep=True)
    df_factor_analysis = df_factor_analysis.set_index("name")
    logger.info(f"df_factor_analysis={df_factor_analysis}")

    df_feature_importance = regression_feature_importance_lightgbm(X_fillna, y_fillna, type_variables, verbose)
    if df_feature_importance.max() < 1:
        df_feature_importance = df_feature_importance * 100
    df_factor_analysis[f"feature_importance"] = df_feature_importance.round(3)
    df_factor_analysis = df_factor_analysis.sort_values(
        ["feature_importance"], ascending=False
    )
    if xlsx_path is not None:
        df_factor_analysis = df_factor_analysis.round(3)
        # df_factor_analysis.to_excel(xlsx_path)
        to_color_excel(df_factor_analysis, xlsx_path)
    return df_factor_analysis


def preprocessing(X_fillna, y_fillna, type_variables, verbose=True):
    from sklearn.preprocessing import LabelEncoder
    from sklearn.feature_extraction.text import TfidfVectorizer
    X_factors = X_fillna.columns.tolist()
    categorical_features_names = [c for c in type_variables.get('Categorical', []) if c in X_factors]
    text_features_names = [c for c in type_variables.get('Text', []) if
                           (c in X_factors) and (c not in categorical_features_names)]
    numeric_features_names = [c for c in type_variables.get('Numeric', []) if
                              (c in X_factors) and (c not in categorical_features_names)]
    logger.info(f'categorical_features_names={categorical_features_names}')
    logger.info(f'text_features_names={text_features_names}')
    logger.info(f'numeric_features_names={numeric_features_names}')

    X_s = [X_fillna[numeric_features_names]]
    if len(categorical_features_names) > 0:
        label_encoder = LabelEncoder()
        # X_categorical = label_encoder.fit_transform(X_fillna[categorical_features_names])
        X_categorical = X_fillna[categorical_features_names].apply(label_encoder.fit_transform)
        X_s.append(X_categorical)
    if len(text_features_names) > 0:
        vectorizer = TfidfVectorizer()
        # X_text = vectorizer.fit_transform(X_fillna[text_features_names])
        X_text = X_fillna[text_features_names].apply(vectorizer.fit_transform)
        X_s.append(X_text)
    if len(X_s) > 0:
        X_transformed = pd.concat(X_s, axis=1)
    else:
        X_transformed = X_s[0]
    return X_transformed


def regression_feature_importance_lightgbm(X_fillna, y_fillna, type_variables, verbose=True) -> pd.Series:
    from lightgbm import LGBMRegressor

    X_factors = X_fillna.columns.tolist()
    X_transformed = preprocessing(X_fillna, y_fillna, type_variables, verbose)

    estimator = LGBMRegressor(n_estimators=2000,
                              feature_fraction=0.06,
                              bagging_fraction=0.67,
                              bagging_freq=1,
                              verbose=int(verbose),
                              random_state=42)
    X_transformed.columns = [f'X{i}' for i in range(len(X_factors))]
    estimator.fit(X_transformed, y_fillna)
    df_feature_importance = pd.Series(
        estimator.feature_importances_, index=X_factors
    )
    return df_feature_importance


def regression_feature_importance_catboost(X_fillna, y_fillna, type_variables, verbose=True) -> pd.Series:
    from catboost import CatBoostRegressor, Pool
    X_factors = X_fillna.columns.tolist()
    categorical_features_names = type_variables.get('Categorical', [])
    categorical_features_names = [c for c in categorical_features_names if c in X_factors]
    text_features_names = type_variables.get('Text', [])
    text_features_names = [c for c in text_features_names if c in X_factors]
    text_features_names = list(set(text_features_names) - set(categorical_features_names))
    logger.info(f'categorical_features_names={categorical_features_names}')
    logger.info(f'text_features_names={text_features_names}')

    # https://github.com/catboost/tutorials/blob/master/text_features/text_features_in_catboost.ipynb

    for f in categorical_features_names:
        if f in type_variables.get('Numeric', []):
            X_fillna[f] = X_fillna[f].astype(int)
    X_text = text_preprocessing(X_fillna, text_features_names)
    train_pool = Pool(data=X_text, label=y_fillna, cat_features=categorical_features_names,
                      text_features=text_features_names,
                      feature_names=list(X_fillna))

    estimator = CatBoostRegressor(
        allow_writing_files=False, verbose=verbose,
        # iterations=1000,  # Number of boosting iterations (trees)
        # learning_rate=0.05,  # Learning rate (shrinkage)
        # depth=6,  # Maximum tree depth
        # l2_leaf_reg=3,  # L2 regularization coefficient
        # random_seed=42,  # Random seed for reproducibility
        # loss_function='RMSE',  # Loss function to optimize
        # eval_metric='RMSE',  # Metric used for validation
        # od_type='Iter',  # Overfitting detection type
        # od_wait=20,  # Number of iterations to wait for overfitting detection
        # verbose=True,  # Frequency of printing progress messages,
    )

    logger.info(f'X_fillna [{X_fillna.shape}], y_fillna [{y_fillna.shape}]')
    estimator.fit(train_pool)
    df_feature_importance = pd.Series(
        estimator.feature_importances_, index=X_factors
    )
    return df_feature_importance


########################################################################################################################
#
########################################################################################################################


def fit_classification(
        df_data: pd.DataFrame,
        X_factors,
        y_factor,
        xlsx_path=None,
        binning_fit_params=None,
        type_variables=None,
        verbose=False,
):
    """
    analyze factor importance of classification task

    Parameters
    ----------
        df_data : pandas.DataFrame
            training table in DataFrame
        X_factors : list[str]
            X factor variable names in df_data columns
        y_factor : str
            y factor variable name in df_data columns
        xlsx_path : str
            result DataFrame output as Excel
        binning_fit_params:dict
            binning fit parameters, see https://gnpalencia.org/optbinning/binning_process.html
        verbose : bool
            logs information output
    Returns
    -------
    df_factor_analysis : pandas.DataFrame
        result table in pandas DataFrame with one factor in each row, including.
        1. dtype: see https://gnpalencia.org/optbinning/binning_tables.html
        2. n_bins: see https://gnpalencia.org/optbinning/binning_tables.html
        3. monotonic_trend: see https://gnpalencia.org/optbinning/binning_tables.html
        4. min_value
        5. max_value
        6. event_rates: see https://gnpalencia.org/optbinning/binning_tables.html
        7. bins: see https://gnpalencia.org/optbinning/binning_tables.html
        8. missing
        9. std
        10. iv: see https://gnpalencia.org/optbinning/binning_tables.html
        11. js: see https://gnpalencia.org/optbinning/binning_tables.html
        12. gini: see https://gnpalencia.org/optbinning/binning_tables.html
        13. quality_score: see https://gnpalencia.org/optbinning/binning_tables.html
        14. feature_importance(RandomForestClassifier): see https://scikit-learn.org/stable/modules/permutation_importance.html
        15. feature_importance(XGBClassifier): see https://scikit-learn.org/stable/modules/permutation_importance.html

    """
    if type_variables is None:
        type_variables = get_data_types(df_data)
    variable_types = reverse_mapping(type_variables)
    X = df_data[X_factors]
    y = df_data[y_factor]
    if verbose:
        logger.info(f"classification factor analysis, with X[{X.shape}], y[{y.shape}], nClass[{y.nunique()}]")
    X_fillna, y_fillna = fill_null_by_types(X, y, variable_types)
    if binning_fit_params is None:
        binning_fit_params = {}
    binning_fit_params_all = dict(
        zip(X_factors, [binning_fit_params] * len(X_factors))
    )
    binning_process = BinningProcess(
        variable_names=X_factors, binning_fit_params=binning_fit_params_all, verbose=verbose
    )
    binning_process.fit(X_fillna, y_fillna)
    df_quality = binning_process.summary()
    if verbose:
        logger.debug(df_quality)
    trends = []
    mins = []
    maxs = []
    event_rates = []
    bins = []
    missings = []
    stds = []
    dtypes = []
    for idx, row in df_quality.iterrows():
        name = row["name"]
        dtype = row["dtype"]
        dtypes.append(dtype)
        if verbose:
            logger.info(f"analyze factor: {name}")
        binning_table = binning_process._binned_variables[name]._binning_table
        if isinstance(binning_table, MulticlassBinningTable):
            mono_strings = []
            for i, c in enumerate(binning_table.classes):
                type_mono = type_of_monotonic_trend(binning_table._event_rate[:-2, i])
                mono_strings.append(f"Class[{i}]]:{c}={type_mono}")
            trend = ' , '.join(mono_strings)
            event_rate_strings = []
            for i, c in enumerate(binning_table.classes):
                event_rate = binning_table._event_rate[:-2, i]
                event_rate_strings.append(f"Class[{i}]]:{c}={event_rate}")
            event_rates.append(' , '.join(event_rate_strings))
        else:
            trend = type_of_monotonic_trend(binning_table._event_rate[:-2])
            event_rates.append(list(binning_table._event_rate[:-2]))
        bins.append(list(binning_table._bin_str[:-2]))
        mins.append(X[name].min())
        maxs.append(X[name].max())
        trends.append(trend)
        missings.append(round(df_data[name].isnull().mean(), 2))

        if dtype in ['numerical']:
            stds.append(df_data[name].std())
        else:
            stds.append(0)

    df_quality["monotonic_trend"] = trends
    df_quality["min_value"] = mins
    df_quality["max_value"] = maxs
    df_quality["event_rates"] = event_rates
    df_quality["bins"] = bins
    df_quality["missing"] = missings
    df_quality["std"] = stds
    df_quality["dtype"] = dtypes

    # columns = ['name', 'dtype',	'status', 'selected', 'n_bins',	'iv',	'js',	'gini',
    #            'quality_score', 'monotonic_trend', 'min_value',	'max_value',
    #            'event_rates', 'bins', 'missing']
    columns = [
        "name",
        "dtype",
        "n_bins",
        "monotonic_trend",
        "min_value",
        "max_value",
        "event_rates",
        "bins",
        "missing",
        "std",
        "iv",
        "js",
        "gini",
        "quality_score",
    ]
    columns = [c for c in columns if c in df_quality.columns]
    df_quality = df_quality[columns]
    df_quality = df_quality.sort_values(["quality_score", "js"], ascending=False)

    df_factor_analysis = df_quality.copy(deep=True)
    df_factor_analysis = df_factor_analysis.set_index("name")

    df_feature_importance = classification_feature_importance_lightgbm(X_fillna, y_fillna, type_variables, verbose)
    if df_feature_importance.max() < 1:
        df_feature_importance = df_feature_importance * 100
    df_factor_analysis[f"feature_importance"] = df_feature_importance.round(3)
    df_factor_analysis = df_factor_analysis.sort_values(
        ["feature_importance"], ascending=False
    )
    if xlsx_path is not None:
        df_factor_analysis = df_factor_analysis.round(3)
        to_color_excel(df_factor_analysis, xlsx_path)
    return df_factor_analysis


def classification_feature_importance_lightgbm(X_fillna, y_fillna, type_variables, verbose=True) -> pd.Series:
    from lightgbm import LGBMClassifier
    from sklearn.preprocessing import LabelEncoder
    from sklearn.feature_extraction.text import TfidfVectorizer

    X_factors = X_fillna.columns.tolist()
    X_transformed = preprocessing(X_fillna, y_fillna, type_variables, verbose)

    # categorical_features_names = [c for c in type_variables.get('Categorical', []) if c in X_factors]
    # text_features_names = [c for c in type_variables.get('Text', []) if
    #                        (c in X_factors) and (c not in categorical_features_names)]
    # numeric_features_names = [c for c in type_variables.get('Numeric', []) if
    #                           (c in X_factors) and (c not in categorical_features_names)]
    # logger.info(f'categorical_features_names={categorical_features_names}')
    # logger.info(f'text_features_names={text_features_names}')
    # logger.info(f'numeric_features_names={numeric_features_names}')
    # X_s = [X_fillna[numeric_features_names]]
    # if len(categorical_features_names) > 0:
    #     label_encoder = LabelEncoder()
    #     X_categorical = label_encoder.fit_transform(X_fillna[categorical_features_names])
    #     X_s.append(X_categorical)
    # if len(text_features_names) > 0:
    #     vectorizer = TfidfVectorizer()
    #     X_text = vectorizer.fit_transform(X_fillna[text_features_names])
    #     X_s.append(X_text)
    # if len(X_s) > 0:
    #     X_transformed = pd.concat(X_s, axis=1)
    # else:
    #     X_transformed = X_s[0]

    estimator = LGBMClassifier(n_estimators=2000,
                               feature_fraction=0.06,
                               bagging_fraction=0.67,
                               bagging_freq=1,
                               verbose=int(verbose),
                               random_state=42)
    X_transformed.columns = [f'X{i}' for i in range(len(X_factors))]
    estimator.fit(X_transformed, y_fillna)
    df_feature_importance = pd.Series(
        estimator.feature_importances_, index=X_factors
    )
    return df_feature_importance


def classification_feature_importance_catboost(X_fillna, y_fillna, type_variables, verbose=True) -> pd.Series:
    from catboost import CatBoostClassifier, Pool

    X_factors = X_fillna.columns.tolist()
    categorical_features_names = type_variables.get('Categorical', [])
    categorical_features_names = [c for c in categorical_features_names if c in X_factors]
    text_features_names = type_variables.get('Text', [])
    text_features_names = [c for c in text_features_names if c in X_factors]
    text_features_names = list(set(text_features_names) - set(categorical_features_names))
    logger.info(f'categorical_features_names={categorical_features_names}')
    logger.info(f'text_features_names={text_features_names}')

    for f in categorical_features_names:
        if f in type_variables.get('Numeric', []):
            X_fillna[f] = X_fillna[f].astype(int)
    X_text = text_preprocessing(X_fillna, text_features_names)
    train_pool = Pool(data=X_text, label=y_fillna, cat_features=categorical_features_names,
                      text_features=text_features_names)
    estimator = CatBoostClassifier(
        allow_writing_files=False, verbose=verbose,
        # iterations=1000,  # Number of boosting iterations (trees)
        # learning_rate=0.05,  # Learning rate (shrinkage)
        # depth=6,  # Maximum tree depth
        # l2_leaf_reg=3,  # L2 regularization coefficient
        # random_seed=42,  # Random seed for reproducibility
        # loss_function='RMSE',  # Loss function to optimize
        # eval_metric='RMSE',  # Metric used for validation
        # od_type='Iter',  # Overfitting detection type
        # od_wait=20,  # Number of iterations to wait for overfitting detection
        # verbose=True  # Frequency of printing progress messages
    )
    estimator.fit(train_pool)
    df_feature_importance = pd.Series(
        estimator.feature_importances_, index=X_factors
    )
    return df_feature_importance


def guess_X_factors(data_types):
    return list(set(data_types.get('Numeric', []) + data_types.get('Categorical', []) + data_types.get('Text', [])))


def guess_y_factors(data_types):
    return list(set(data_types.get('Numeric', []) + data_types.get('Categorical', []) + data_types.get('Text', [])))
